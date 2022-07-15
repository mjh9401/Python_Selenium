from openpyxl import load_workbook   # openpyxl 중에서 load_workbook 함수만 사용


# 읽을 엑셀 파일
xl_file = '부동산공시가격.xlsx'

# 엑셀 파일 불러오기
wb = load_workbook(filename=xl_file)

# 시트 개수 확인
print('sheet count: ',len(wb.sheetnames))

# 시트별 루프
for sheet in wb.worksheets:
    # 시트 이름 출력
    print('[sheet name : {}]'.format(sheet.title))

    # 한 줄씩(행,row) 읽음
    for row in sheet.rows:
        # 행에서 한 개의 셀(cell) 단위로 인덱스와 함께 읽는다.
        for idx,cell in enumerate(row):
            # 처음을 제외하고 각 셀을 구분하기 위해 탭을 넣어준다.
            if idx != 0:
                print('\t',end='')

            # 값이 없는(None) 경우 무시한다.
            if cell.value is None:
                continue

            # 셀이 있는 값을 출력한다.
            print(cell.value,end='')

        # 줄 바꿈 각 행이 끝나면 줄 바꿈
        print('')

    # 줄 바꿈 각 시트가 끝나면 줄 바꿈
    print('') 
# 완료 
print('Done')  